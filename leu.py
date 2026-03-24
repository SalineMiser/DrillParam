"""
Drilling Real-Time Log Viewer
==============================
• Upload Excel (.xlsx) or CSV/TXT — any format, first row = headers, second row = units (auto-skipped)
• Mud-log style vertical strip charts:  Y = Time or Bit Depth,  X = parameter value
• Uploaded files are stored permanently on the server — reload any time
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import re, os, io, json
from datetime import datetime
import openpyxl

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Drilling Log Viewer", page_icon=None,
                   layout="wide", initial_sidebar_state="expanded")

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] {
    font-family: "Inter", "Segoe UI", system-ui, sans-serif;
    letter-spacing: -0.01em;
}
.stApp { background: #f0f4f8; }
[data-testid="stSidebar"] { background: #ffffff; border-right: 1px solid #d1dbe8; }
[data-testid="stSidebar"] * { color: #344055 !important; }
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    color: #1a2535 !important; font-weight: 600 !important;
    font-size: .85rem !important; text-transform: uppercase; letter-spacing: .06em;
}
[data-testid="stMetric"] {
    background: #ffffff; border: 1px solid #d1dbe8; border-radius: 8px;
    padding: 14px 18px; box-shadow: 0 1px 3px rgba(0,0,0,.04);
}
[data-testid="stMetricValue"] {
    color: #1558a0 !important; font-size: 1.3rem !important;
    font-weight: 700 !important; letter-spacing: -.02em;
}
[data-testid="stMetricLabel"] {
    color: #64748b !important; font-size: .72rem !important;
    text-transform: uppercase; letter-spacing: .05em; font-weight: 500 !important;
}
.top-bar {
    display: flex; align-items: center; justify-content: space-between;
    background: #ffffff; border: 1px solid #d1dbe8; border-radius: 8px;
    padding: 14px 24px; margin-bottom: 20px; box-shadow: 0 1px 4px rgba(0,0,0,.04);
}
.top-bar-title { font-size: 1.1rem; font-weight: 700; color: #1a2535; letter-spacing: -.02em; }
.sec {
    color: #1a2535; font-size: .78rem; font-weight: 600; text-transform: uppercase;
    letter-spacing: .07em; padding-bottom: 6px; border-bottom: 1.5px solid #d1dbe8;
    margin-bottom: 12px;
}
/* Y-axis mode toggle pill */
.ymode-bar {
    display: inline-flex; border: 1px solid #d1dbe8; border-radius: 6px;
    overflow: hidden; margin-bottom: 12px;
}
.ymode-btn {
    padding: 6px 18px; font-size: .78rem; font-weight: 600; cursor: pointer;
    text-transform: uppercase; letter-spacing: .05em; transition: background .15s;
    background: #ffffff; color: #64748b; border: none;
}
.ymode-btn.active { background: #1558a0; color: #ffffff; }
[data-testid="stExpander"] {
    background: #ffffff; border: 1px solid #d1dbe8 !important; border-radius: 8px;
}
.stTabs [data-baseweb="tab-list"] {
    background: #ffffff; border-radius: 8px; border: 1px solid #d1dbe8; gap: 0;
}
.stTabs [data-baseweb="tab"] {
    color: #64748b; font-size: .8rem; font-weight: 500;
    text-transform: uppercase; letter-spacing: .04em; padding: 8px 20px;
}
.stTabs [aria-selected="true"] {
    color: #1558a0 !important; border-bottom: 2px solid #1558a0 !important; font-weight: 600 !important;
}
.stButton > button { font-size: .8rem; font-weight: 500; letter-spacing: .02em; border-radius: 6px; }
[data-testid="stDownloadButton"] button { font-size: .78rem; font-weight: 500; letter-spacing: .02em; }
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #c0ccd8; border-radius: 4px; }
</style>""", unsafe_allow_html=True)

# ── Activity colours ───────────────────────────────────────────────────────────
ACT = {
    "Drilling":       "#16a34a",
    "Tripping In":    "#2563eb",
    "Tripping Out":   "#0891b2",
    "Connection":     "#d97706",
    "Hoisting":       "#7c3aed",
    "Slips / Static": "#64748b",
}

PALETTE = ["#e63946","#2a9d8f","#e76f51","#457b9d","#f4a261",
           "#6a4c93","#1d3557","#2b9348","#c77dff","#ff6b6b",
           "#06d6a0","#ffd166","#118ab2","#ef476f","#073b4c"]

# ═══════════════════════════════════════════════════════════════════════════════
#  FILE STORAGE
# ═══════════════════════════════════════════════════════════════════════════════
STORE_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploaded_files")
INDEX_FILE = os.path.join(STORE_DIR, "index.json")
os.makedirs(STORE_DIR, exist_ok=True)

def _idx():
    return json.load(open(INDEX_FILE)) if os.path.exists(INDEX_FILE) else {}

def _save_idx(d):
    json.dump(d, open(INDEX_FILE, "w"), indent=2)

def store_file(name: str, data: bytes) -> str:
    safe = re.sub(r"[^\w.\-]", "_", name)
    open(os.path.join(STORE_DIR, safe), "wb").write(data)
    idx = _idx()
    idx[safe] = {"original": name, "size": len(data), "saved": datetime.now().isoformat()[:19]}
    _save_idx(idx)
    return safe

def list_stored():
    return [(k, v) for k, v in sorted(_idx().items(), key=lambda x: x[1].get("saved", ""), reverse=True)
            if os.path.exists(os.path.join(STORE_DIR, k))]

def del_stored(safe):
    p = os.path.join(STORE_DIR, safe)
    if os.path.exists(p): os.remove(p)
    idx = _idx(); idx.pop(safe, None); _save_idx(idx)

def read_stored(safe):
    return open(os.path.join(STORE_DIR, safe), "rb").read()

# ═══════════════════════════════════════════════════════════════════════════════
#  SMART UNIVERSAL PARSER
# ═══════════════════════════════════════════════════════════════════════════════
def _looks_like_units(row):
    vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
    if not vals: return True
    unit_like = sum(1 for v in vals
                    if re.match(r'^[a-zA-Z/%°³\-–—_]+$', v) or v in ("None", ""))
    return unit_like / len(vals) >= 0.5

def _parse_time(v):
    if v is None: return pd.NaT
    if isinstance(v, datetime): return pd.Timestamp(v)
    try: return pd.to_datetime(str(v), dayfirst=True)
    except: return pd.NaT

def smart_parse(data: bytes, filename: str):
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        sh = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
        all_rows = list(sh.iter_rows(values_only=True))
    else:
        text = data.decode("utf-8", errors="ignore")
        lines = [l for l in text.splitlines() if l.strip()]
        first = lines[0]
        sep = "\t" if first.count("\t") >= 2 else (";" if first.count(";") >= 2 else ",")
        all_rows = [[c.strip().strip('"') for c in l.split(sep)] for l in lines]

    if not all_rows:
        raise ValueError("File appears to be empty.")

    headers = [str(h).strip() if h is not None else f"Col_{i}"
               for i, h in enumerate(all_rows[0])]
    while headers and headers[-1] in ("None", "", "Col_" + str(len(headers) - 1)):
        headers.pop()
    n_cols = len(headers)

    data_start = 1
    detected_units = {}
    if len(all_rows) > 1 and _looks_like_units(all_rows[1][:n_cols]):
        data_start = 2
        for i, h in enumerate(headers):
            uval = all_rows[1][i] if i < len(all_rows[1]) else None
            if uval is not None and str(uval).strip() not in ("None", ""):
                detected_units[h] = str(uval).strip()

    rows = [list(r)[:n_cols] for r in all_rows[data_start:] if any(v is not None for v in r)]
    df = pd.DataFrame(rows, columns=headers)

    time_col = headers[0]
    for h in headers:
        if re.search(r'time|date|rig', h, re.I):
            time_col = h; break

    df["_rt_parsed"] = df[time_col].apply(_parse_time)
    df = df.dropna(subset=["_rt_parsed"]).reset_index(drop=True)

    for c in df.columns:
        if c in ("_rt_parsed", time_col): continue
        df[c] = pd.to_numeric(df[c], errors="coerce")

    if time_col != "_rt_parsed":
        df = df.drop(columns=[time_col], errors="ignore")
    df = df.dropna(axis=1, how="all")
    df = df.rename(columns={"_rt_parsed": "RigTime"})
    df = df.sort_values("RigTime").drop_duplicates("RigTime").reset_index(drop=True)

    if len(df) < 2:
        raise ValueError("Fewer than 2 valid rows after parsing — check your file.")
    return df, detected_units

def find_col(df, *patterns):
    for p in patterns:
        for c in df.columns:
            if re.search(p, c, re.I): return c
    return None

# ═══════════════════════════════════════════════════════════════════════════════
#  ACTIVITY
# ═══════════════════════════════════════════════════════════════════════════════
def classify(hl, dd, hl_trip, hl_conn, d_thr):
    if hl > hl_trip:
        if dd < -0.3: return "Tripping In"
        if dd >  0.3: return "Tripping Out"
        return "Hoisting"
    if hl > hl_conn: return "Connection"
    if dd < -d_thr:  return "Drilling"
    return "Slips / Static"

def add_activity(df, hl_col, depth_col, hl_trip, hl_conn, d_thr):
    acts = []
    for i in range(len(df)):
        dd = df[depth_col].iloc[i] - df[depth_col].iloc[i - 1] if i > 0 else 0
        hl = float(df[hl_col].iloc[i]) if hl_col else 0
        acts.append(classify(hl, dd, hl_trip, hl_conn, d_thr))
    df = df.copy(); df["_Activity"] = acts; return df

def detect_events(df, depth_col):
    recs, prev, s = [], df["_Activity"].iloc[0], 0
    for i in range(1, len(df)):
        if df["_Activity"].iloc[i] != prev:
            recs.append({"activity": prev,
                "start_time": df["RigTime"].iloc[s], "end_time": df["RigTime"].iloc[i - 1],
                "duration_min": (df["RigTime"].iloc[i - 1] - df["RigTime"].iloc[s]).total_seconds() / 60,
                "depth_start": df[depth_col].iloc[s], "depth_end": df[depth_col].iloc[i - 1]})
            prev, s = df["_Activity"].iloc[i], i
    recs.append({"activity": prev,
        "start_time": df["RigTime"].iloc[s], "end_time": df["RigTime"].iloc[-1],
        "duration_min": (df["RigTime"].iloc[-1] - df["RigTime"].iloc[s]).total_seconds() / 60,
        "depth_start": df[depth_col].iloc[s], "depth_end": df[depth_col].iloc[-1]})
    return pd.DataFrame(recs)

# ═══════════════════════════════════════════════════════════════════════════════
#  MUD LOG — vertical strip chart
#
#  y_mode = "time"  →  Y axis = Rig Time  (oldest at top)
#  y_mode = "depth" →  Y axis = Bit Depth (shallow at top, deep at bottom)
#
#  When y_mode = "depth":
#   • The dedicated Bit Depth strip is hidden (it IS the Y axis)
#   • Activity bands are drawn as horizontal depth-interval rectangles
#   • All param traces use depth on Y instead of time
# ═══════════════════════════════════════════════════════════════════════════════
def build_mud_log(df, param_cols, depth_col, events_df,
                  show_bands, height, ds_n, units, y_mode="time"):

    if units is None:
        units = {}

    use_depth_y = (y_mode == "depth")

    # In depth mode the dedicated depth strip is replaced by the Y axis itself
    n          = len(param_cols)
    total_cols = n if use_depth_y else (1 + n)   # depth strip only in time mode

    step = max(1, len(df) // ds_n)
    ds   = df.iloc[::step].copy()

    if use_depth_y:
        widths = [1.0] * n
    else:
        widths = [0.85] + [1.0] * n   # depth strip slightly narrower on left

    fig = make_subplots(
        rows=1, cols=total_cols,
        shared_yaxes=True,
        column_widths=widths,
        horizontal_spacing=0.008,
    )

    # ── Shared Y values ───────────────────────────────────────────────────────
    if use_depth_y:
        y_vals     = ds[depth_col]
        y_min      = float(ds[depth_col].min())
        y_max      = float(ds[depth_col].max())
        y_range    = [y_min - 2, y_max + 2]   # shallow at top → NOT reversed
        y_reversed = False
        y_fmt      = None
        y_title    = f"{depth_col} ({units.get(depth_col,'m')})"
    else:
        y_vals     = ds["RigTime"]
        y_range    = None
        y_reversed = True                      # oldest at top
        y_fmt      = "%H:%M\n%d %b"
        y_title    = "TIME"

    # ── Activity bands ────────────────────────────────────────────────────────
    if show_bands and events_df is not None:
        for _, ev in events_df.iterrows():
            band_color = ACT.get(ev["activity"], "#64748b")
            if use_depth_y:
                # horizontal bands in depth space
                y0 = min(ev["depth_start"], ev["depth_end"])
                y1 = max(ev["depth_start"], ev["depth_end"])
                if y1 - y0 < 0.1: y0 -= 0.2; y1 += 0.2
            else:
                y0 = ev["start_time"]
                y1 = ev["end_time"]
            for ci in range(1, total_cols + 1):
                fig.add_hrect(y0=y0, y1=y1,
                              fillcolor=band_color, opacity=0.07, line_width=0,
                              row=1, col=ci)

    # ── Helper: strip header label ────────────────────────────────────────────
    def _label(col_name, color=None):
        u    = units.get(col_name, "").strip()
        unit = f" ({u})" if u else ""
        name = f"{col_name}{unit}"
        if color:
            return f"<b style='color:{color}'>{name}</b>"
        return f"<b>{name}</b>"

    # ── Col 1 in TIME mode: Bit Depth strip ───────────────────────────────────
    if not use_depth_y:
        d_series = ds[depth_col].dropna()
        if not d_series.empty:
            d_min = float(d_series.min()); d_max = float(d_series.max())
            d_unit = units.get(depth_col, "").strip()
            hu = f" {d_unit}" if d_unit else ""

            fig.add_trace(go.Scatter(
                x=ds[depth_col], y=ds["RigTime"],
                mode="lines",
                name=_label(depth_col).replace("<b>", "").replace("</b>", ""),
                line=dict(color="#1a2535", width=2.5),
                fill="tozerox", fillcolor="rgba(26,37,53,0.06)",
                hovertemplate=(f"<b>{depth_col}</b><br>"
                               f"%{{y|%H:%M:%S}}<br>%{{x:.2f}}{hu}<extra></extra>"),
            ), row=1, col=1)

            fig.update_xaxes(
                range=[d_min - 5, d_max + 5],
                tickfont=dict(size=8, color="#94a3b8"),
                gridcolor="#e8eef4", gridwidth=1,
                linecolor="#d1dbe8", zeroline=False,
                showgrid=True, side="top",
                title_text=(_label(depth_col) + "<br>"
                            f"<span style='font-size:9px;color:#94a3b8'>"
                            f"{d_min:.0f} … {d_max:.0f}</span>"),
                title_font=dict(size=10, color="#1a2535"),
                title_standoff=2,
                row=1, col=1,
            )

    # ── Parameter strips ──────────────────────────────────────────────────────
    for i, col in enumerate(param_cols):
        color  = PALETTE[i % len(PALETTE)]
        ci     = i + 1 if use_depth_y else i + 2
        series = ds[col].dropna()
        if series.empty: continue

        xmin = float(ds[col].min()); xmax = float(ds[col].max())
        pad  = max((xmax - xmin) * 0.04, 0.5)
        u    = units.get(col, "").strip()
        hu   = f" {u}" if u else ""

        if use_depth_y:
            y_data = ds[depth_col]
            hover  = (f"<b>{col}</b><br>"
                      f"Depth: %{{y:.2f}} m<br>Value: %{{x:.2f}}{hu}<extra></extra>")
        else:
            y_data = ds["RigTime"]
            hover  = (f"<b>{col}</b><br>"
                      f"%{{y|%H:%M:%S}}<br>%{{x:.2f}}{hu}<extra></extra>")

        fig.add_trace(go.Scatter(
            x=ds[col], y=y_data,
            mode="lines",
            name=f"{col}{hu}",
            line=dict(color=color, width=1.3),
            fill="tozerox",
            fillcolor=f"rgba({int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)},0.08)",
            hovertemplate=hover,
        ), row=1, col=ci)

        fig.update_xaxes(
            range=[xmin - pad, xmax + pad],
            tickfont=dict(size=8, color="#94a3b8"),
            gridcolor="#e8eef4", gridwidth=1,
            linecolor="#d1dbe8", zeroline=False,
            showgrid=True, side="top",
            title_text=(_label(col, color) + "<br>"
                        f"<span style='color:#94a3b8;font-size:9px'>"
                        f"{xmin:.1f} … {xmax:.1f}</span>"),
            title_font=dict(size=10),
            title_standoff=2,
            row=1, col=ci,
        )

    # ── Shared Y axis ─────────────────────────────────────────────────────────
    yax = dict(
        tickfont=dict(size=9, color="#64748b"),
        gridcolor="#e8eef4", gridwidth=1,
        linecolor="#d1dbe8", zeroline=False,
        showgrid=True,
        title_text=y_title,
        title_font=dict(size=10, color="#374151"),
    )
    if y_fmt:
        yax["tickformat"] = y_fmt
    if y_range:
        yax["range"]      = y_range
        yax["autorange"]  = False
    else:
        yax["autorange"]  = "reversed" if y_reversed else True

    fig.update_yaxes(**yax, row=1, col=1)
    for ci in range(2, total_cols + 1):
        fig.update_yaxes(showticklabels=False, row=1, col=ci)

    fig.update_layout(
        height=height,
        paper_bgcolor="#ffffff",
        plot_bgcolor="#fafbfd",
        font=dict(family="Inter, sans-serif", color="#374151"),
        legend=dict(
            orientation="h", yanchor="bottom", y=1.04, xanchor="left", x=0,
            font=dict(size=10), bgcolor="rgba(255,255,255,0.92)",
            bordercolor="#d1dbe8", borderwidth=1,
        ),
        margin=dict(l=65, r=15, t=90, b=30),
        hovermode="y unified",
        hoverlabel=dict(bgcolor="#fff", bordercolor="#d1dbe8",
                        font=dict(color="#1a2535", size=11)),
    )
    return fig

# ═══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
<div style='padding:18px 4px 14px 4px;border-bottom:1px solid #d1dbe8;margin-bottom:6px;'>
  <div style='font-size:.65rem;font-weight:600;color:#94a3b8;text-transform:uppercase;
              letter-spacing:.08em;margin-bottom:4px;'>Engineer</div>
  <div style='font-size:1rem;font-weight:700;color:#1558a0;letter-spacing:-.01em;'>
              Saleh Aliyev</div>
  <div style='font-size:.72rem;color:#94a3b8;margin-top:2px;letter-spacing:.01em;'>
              Drilling Real-Time Log Viewer</div>
</div>""", unsafe_allow_html=True)

    st.markdown("### Display")
    chart_height = st.slider("Chart height (px)", 600, 3000, 1500, 50)
    ds_n         = st.slider("Max points",         500, 8000, 3000, 100)
    show_bands   = st.toggle("Activity bands", True)

    st.markdown("---")
    st.markdown("### Activity Thresholds")
    hl_trip = st.number_input("Tripping HL (t)",      value=35.0, step=0.5)
    hl_conn = st.number_input("Connection HL (t)",    value=10.0, step=0.5)
    dd_thr  = st.number_input("Drilling depth Δ (m)", value=0.05, step=0.01)

# ═══════════════════════════════════════════════════════════════════════════════
#  HEADER
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="top-bar">
  <div class="top-bar-left">
    <span class="top-bar-title">Drilling Real-Time Log Viewer</span>
  </div>
</div>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  FILE MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════
tab_up, tab_stored = st.tabs(["Upload File", "Stored Files"])

active_bytes = None
active_name  = None

with tab_up:
    st.caption("**Excel (.xlsx)** or **CSV / TXT**. "
               "Row 1 = parameter names. Row 2 = units if present — detected and skipped automatically.")
    uf = st.file_uploader("", type=["xlsx", "xls", "csv", "txt", "tsv"],
                          label_visibility="collapsed")
    if uf:
        raw = uf.read()
        store_file(uf.name, raw)
        active_bytes, active_name = raw, uf.name
        st.success(f"**{uf.name}** saved to server storage.")

with tab_stored:
    files = list_stored()
    if not files:
        st.info("No files stored yet. Upload a file using the tab above.")
    else:
        st.caption(f"{len(files)} file(s) stored on server — click a row to load, or delete.")
        for safe, meta in files:
            ca, cb = st.columns([6, 1])
            with ca:
                label = (f"**{meta['original']}**  ·  "
                         f"{meta['saved'][:16]}  ·  {meta['size'] // 1024} KB")
                if st.button(label, key=f"ld_{safe}", use_container_width=True):
                    active_bytes = read_stored(safe)
                    active_name  = meta["original"]
                    st.success(f"Loaded: **{meta['original']}**")
            with cb:
                if st.button("×", key=f"dl_{safe}", help="Delete this file"):
                    del_stored(safe); st.rerun()

if active_bytes is None:
    files = list_stored()
    if files:
        safe, meta = files[0]
        active_bytes = read_stored(safe)
        active_name  = meta["original"]
        st.info(f"Auto-loaded most recent file: **{meta['original']}**")
    else:
        st.warning("Upload a file to get started.")
        st.stop()

# ═══════════════════════════════════════════════════════════════════════════════
#  PARSE
# ═══════════════════════════════════════════════════════════════════════════════
try:
    df, auto_units = smart_parse(active_bytes, active_name)
except Exception as e:
    st.error(f"**Parse error:** {e}")
    st.stop()

num_cols  = [c for c in df.columns
             if c != "RigTime" and pd.api.types.is_numeric_dtype(df[c])]
depth_col = find_col(df, r'bitdep', r'depth', r'\bmd\b') or (num_cols[0] if num_cols else None)
hl_col    = find_col(df, r'hook.?load', r'\bhl\b')

# ═══════════════════════════════════════════════════════════════════════════════
#  COLUMN SETUP
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("<div class='sec'>Column Setup</div>", unsafe_allow_html=True)
cc1, cc2, cc3 = st.columns(3)

with cc1:
    depth_col = st.selectbox("Bit Depth column",
                             num_cols, index=num_cols.index(depth_col) if depth_col in num_cols else 0)
with cc2:
    hl_opts = ["(none)"] + num_cols
    hl_col  = st.selectbox("Hook Load (for activity)",
                           hl_opts, index=hl_opts.index(hl_col) if hl_col in hl_opts else 0)
    if hl_col == "(none)": hl_col = None
with cc3:
    avail      = [c for c in num_cols if c != depth_col]
    default    = avail[:min(6, len(avail))]
    param_cols = st.multiselect("Parameter tracks (each = one strip)", avail, default=default)

if not param_cols:
    st.warning("Pick at least one parameter track above."); st.stop()
if depth_col is None:
    st.warning("Could not find a Bit Depth column."); st.stop()

# ── Units editor ───────────────────────────────────────────────────────────────
with st.expander("Units — optional, shown in strip headers", expanded=False):
    st.caption("Units were auto-detected from the file where available. Override any value here.")
    all_track_cols = [depth_col] + param_cols
    n_ucols = min(4, len(all_track_cols))
    ucols = st.columns(n_ucols)
    units = {}
    for idx, col in enumerate(all_track_cols):
        default_unit = auto_units.get(col, "")
        with ucols[idx % n_ucols]:
            units[col] = st.text_input(col, value=default_unit, key=f"unit_{col}",
                                       placeholder="e.g. m, t, rpm")

# ═══════════════════════════════════════════════════════════════════════════════
#  Y-AXIS MODE SELECTOR
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("<div class='sec'>Chart Y Axis</div>", unsafe_allow_html=True)

y_col1, y_col2 = st.columns([2, 5])
with y_col1:
    y_mode = st.radio(
        "Y axis",
        options=["time", "depth"],
        format_func=lambda x: "Rig Time" if x == "time" else "Bit Depth",
        horizontal=True,
        label_visibility="collapsed",
    )
with y_col2:
    if y_mode == "time":
        st.caption(
            "Strips share a **time axis** — oldest at top, newest at bottom. "
            "The leftmost strip shows Bit Depth as a reference track."
        )
    else:
        st.caption(
            "Strips share a **depth axis** — shallow at top, deeper at bottom. "
            "Bit Depth becomes the Y axis itself; all parameters are plotted against measured depth."
        )

# ═══════════════════════════════════════════════════════════════════════════════
#  ACTIVITY
# ═══════════════════════════════════════════════════════════════════════════════
if hl_col:
    df  = add_activity(df, hl_col, depth_col, hl_trip, hl_conn, dd_thr)
    evs = detect_events(df, depth_col)
else:
    df["_Activity"] = "Slips / Static"
    evs = None

# ═══════════════════════════════════════════════════════════════════════════════
#  METRICS
# ═══════════════════════════════════════════════════════════════════════════════
t_sec  = (df["RigTime"].iloc[-1] - df["RigTime"].iloc[0]).total_seconds()
t_hrs  = t_sec / 3600
drl    = abs(df[depth_col].iloc[-1] - df[depth_col].iloc[0])
rop    = drl / t_hrs if t_hrs > 0 else 0
max_hl = df[hl_col].max() if hl_col else 0
n_conn = len(evs[evs.activity == "Connection"]) if evs is not None else 0
d_min  = df["RigTime"].iloc[0].strftime("%d %b %H:%M")
d_max  = df["RigTime"].iloc[-1].strftime("%d %b %H:%M")

c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("Period",        f"{t_hrs:.1f} hr",   delta=f"{d_min} → {d_max}")
c2.metric("Depth Range",   f"{df[depth_col].min():.0f} – {df[depth_col].max():.0f} m",
                            delta=f"{drl:.1f} m drilled")
c3.metric("Avg ROP",       f"{rop:.2f} m/hr")
c4.metric("Max Hook Load", f"{max_hl:.1f} t")
c5.metric("Connections",   str(n_conn))
c6.metric("Channels",      str(len(param_cols)), delta=f"{len(df):,} rows")

st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  MUD LOG CHART
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("<div class='sec'>Mud Log — Vertical Strip Chart</div>", unsafe_allow_html=True)

fig = build_mud_log(df, param_cols, depth_col, evs,
                    show_bands, chart_height, ds_n, units, y_mode)
st.plotly_chart(fig, use_container_width=True, config={
    "displayModeBar": True, "scrollZoom": True,
    "modeBarButtonsToRemove": ["lasso2d", "select2d"],
    "toImageButtonOptions": {"format": "png", "filename": "mud_log", "scale": 2},
})

# ═══════════════════════════════════════════════════════════════════════════════
#  ACTIVITY BREAKDOWN + EVENT LOG
# ═══════════════════════════════════════════════════════════════════════════════
if evs is not None:
    left, right = st.columns([1, 2.4])

    with left:
        st.markdown("<div class='sec'>Activity Breakdown</div>", unsafe_allow_html=True)
        agg = (evs.groupby("activity")["duration_min"].sum()
               .reset_index().sort_values("duration_min", ascending=False))
        agg["pct"] = (agg["duration_min"] / agg["duration_min"].sum() * 100).round(1)
        total  = agg["duration_min"].sum()
        colors = [ACT.get(a, "#64748b") for a in agg["activity"]]
        donut  = go.Figure(go.Pie(
            labels=agg["activity"], values=agg["duration_min"].round(1),
            hole=0.60, marker=dict(colors=colors, line=dict(color="#f0f4f8", width=3)),
            textinfo="percent", textfont=dict(size=11),
            hovertemplate="<b>%{label}</b><br>%{value:.1f} min (%{percent})<extra></extra>",
        ))
        donut.update_layout(
            annotations=[dict(text=f"<b>{total:.0f}</b><br>min",
                              x=0.5, y=0.5, font_size=15, font_color="#1a2535", showarrow=False)],
            paper_bgcolor="#ffffff", font=dict(color="#374151"),
            legend=dict(font=dict(size=10), bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=10, r=10, t=10, b=10), height=230,
        )
        st.plotly_chart(donut, use_container_width=True, config={"displayModeBar": False})

        rows_h = ""
        for _, r in agg.iterrows():
            c = ACT.get(r["activity"], "#64748b")
            rows_h += (f"<tr><td style='padding:5px 8px'>"
                       f"<span style='display:inline-block;width:9px;height:9px;border-radius:2px;"
                       f"background:{c};margin-right:6px'></span>{r['activity']}</td>"
                       f"<td style='text-align:right;padding:5px 8px'>{r['duration_min']:.0f} min</td>"
                       f"<td style='text-align:right;padding:5px 8px'>{r['pct']}%</td></tr>")
        st.markdown(
            f"<table style='width:100%;border-collapse:collapse;font-size:.8rem;color:#334155'>"
            f"<thead><tr>"
            f"<th style='text-align:left;padding:5px 8px;border-bottom:2px solid #d1dbe8;"
            f"color:#64748b;font-size:.68rem;text-transform:uppercase'>Activity</th>"
            f"<th style='text-align:right;padding:5px 8px;border-bottom:2px solid #d1dbe8;"
            f"color:#64748b;font-size:.68rem;text-transform:uppercase'>Min</th>"
            f"<th style='text-align:right;padding:5px 8px;border-bottom:2px solid #d1dbe8;"
            f"color:#64748b;font-size:.68rem;text-transform:uppercase'>%</th>"
            f"</tr></thead><tbody>{rows_h}</tbody></table>",
            unsafe_allow_html=True,
        )

    with right:
        st.markdown("<div class='sec'>Event Log</div>", unsafe_allow_html=True)
        badge = {"Drilling": "#dcfce7|#166534", "Tripping In": "#dbeafe|#1e40af",
                 "Tripping Out": "#cffafe|#0e7490", "Hoisting": "#ede9fe|#5b21b6",
                 "Connection": "#fef9c3|#854d0e", "Slips / Static": "#f1f5f9|#475569"}
        rows_h = ""
        for i, ev in evs.iterrows():
            bg, fg = badge.get(ev["activity"], "#f1f5f9|#475569").split("|")
            dd = abs(ev["depth_end"] - ev["depth_start"])
            rows_h += (
                f"<tr style='border-bottom:1px solid #eef2f7'>"
                f"<td style='padding:5px 9px;color:#94a3b8'>{i + 1}</td>"
                f"<td style='padding:5px 9px'>{ev['start_time'].strftime('%d %b %H:%M:%S')}</td>"
                f"<td style='padding:5px 9px'>{ev['end_time'].strftime('%H:%M:%S')}</td>"
                f"<td style='padding:5px 9px'>{ev['duration_min']:.1f} min</td>"
                f"<td style='padding:5px 9px'>"
                f"<span style='background:{bg};color:{fg};padding:2px 9px;border-radius:12px;"
                f"font-size:.73rem;font-weight:600'>{ev['activity']}</span></td>"
                f"<td style='padding:5px 9px;font-size:.8rem'>"
                f"{ev['depth_start']:.1f} → {ev['depth_end']:.1f} m</td>"
                f"<td style='padding:5px 9px;color:#64748b'>{dd:.1f} m</td>"
                f"</tr>"
            )
        st.markdown(
            f"<div style='max-height:370px;overflow-y:auto;border:1px solid #d1dbe8;"
            f"border-radius:10px;background:#fff'>"
            f"<table style='width:100%;border-collapse:collapse;font-size:.8rem;color:#334155'>"
            f"<thead style='position:sticky;top:0;background:#f8fafc'><tr>"
            f"{''.join(f'<th style=padding:6px_9px;border-bottom:2px_solid_#d1dbe8;text-align:left;color:#64748b;font-size:.68rem;text-transform:uppercase>{h}</th>' for h in ['#', 'Start', 'End', 'Duration', 'Activity', 'Depth', 'Delta'])}"
            f"</tr></thead><tbody>{rows_h}</tbody></table></div>",
            unsafe_allow_html=True,
        )
        st.download_button("Download Events CSV", evs.to_csv(index=False).encode(),
                           "events.csv", "text/csv")

# ═══════════════════════════════════════════════════════════════════════════════
#  RAW DATA
# ═══════════════════════════════════════════════════════════════════════════════
with st.expander("Raw Data Preview & Download", expanded=False):
    st.dataframe(df.head(500), use_container_width=True, height=280)
    st.download_button("Download Full CSV", df.to_csv(index=False).encode(),
                       "rig_data.csv", "text/csv")
